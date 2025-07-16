import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from itertools import product

# Define control variables
"""
"temperature": 
Creative writing: High temperature can be beneficial for
generating imaginative stories or poems. Factual responses: Low temperature
is often preferred for tasks requiring precise answers to questions. 

"top_p":
Chooses tokens based on how much probability mass, and represents a cumulative
probability threshold. Lower values: Lead to more focused and deterministic
outputs, as the model only considers a smaller set of high-probability tokens. 
Higher values: Encourage more diverse and creative responses by including
a wider range of potential tokens. 
"""
control_variables = {
    'model': ['TeenyTinyLlama-160m-NCM-ft', 'deepseek-reasoner', 'gpt-4o-mini-2024-07-18', 'o1-mini-2024-09-12', 'Mistral-7B-Instruct-v0.3', 'gemini-2.0-flash'],
    'attempt': range(1, 197),
    'temperature': [0.1, 1.0, 1.9],
    'top_p': [0.1, 0.5, 0.9]
}

# Create a list of all combinations of control variables
keys = control_variables.keys()
values = control_variables.values()
combinations = list(product(*values))

# Create a pandas DataFrame from the combinations
df_final = pd.DataFrame(combinations, columns=keys)

#Rename columns for better understanding
df_final = df_final.rename(columns={'replicates':'replicate'})

# Add a column for the prompt (to be filled later)
df_final['prompt'] = ''

# Add a column for the results (to be filled later)
df_final['results'] = ''

# Add a column for the baseline (to be filled later)
df_final['baseline'] = ''

# Add a column for the score (to be filled later)
df_final['score'] = ''

# Save the DataFrame to an XLSX file with good formatting
wb = Workbook()
ws = wb.active

# Write the header
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
for col_num, value in enumerate(df_final.columns.values):
    ws.cell(row=1, column=col_num+1, value=value).font = header_font
    ws.cell(row=1, column=col_num+1).fill = header_fill

# Write the data
for row in df_final.values.tolist():
    ws.append(row)

wb.save("experimental_design_plan.xlsx")

import numpy as np
from statsmodels.stats.power import FTestAnovaPower

# Experiment parameters:
effect_size = 0.25  # Medium effect size (Cohen's f)
alpha = 0.05        # Significance level
power = 0.80        # Desired power (80%)
k_groups = 5        # Number of groups (e.g., for the 'model' factor)

# Initialize the power analysis object for ANOVA
analysis = FTestAnovaPower()

# Calculate the sample size per group (replicates per condition)
n = analysis.solve_power(effect_size=effect_size, alpha=alpha, power=power, k_groups=k_groups)

print(f"Number of replicates per condition: {np.ceil(n)}")
